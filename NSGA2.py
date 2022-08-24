# -*- coding: utf-8 -*-
import numpy as np
import geatpy as ea
import openpyxl


# =====================================9个水电站的主要函数表======================================
def h1v1_w1(hx):
    vx = round(3.33257987E-06 * float(hx) ** 4 - 3.40473719E-03 * float(hx) ** 3 + 1.30798512E+00 * float(
        hx) ** 2 - 2.23769818E+02 * float(hx) + 1.43757521E+04, 5)
    return round(vx, 5)


def v1h1_w1(vx):
    hx = round(
        -1.8741E-02 * float(vx) ** 4 + 3.5765E-01 * float(vx) ** 3 - 2.7713E+00 * float(vx) ** 2 + 1.3795E+01 * float(
            vx) + 2.3637E+02, 5)
    return round(hx, 5)


def qh2_w1(qx):
    h2x = -1.4364E-17 * float(qx) ** 6 + 7.5688E-14 * float(qx) ** 5 - 1.5723E-10 * float(qx) ** 4 + 1.6373E-07 * float(
        qx) ** 3 - 9.0046E-05 * float(qx) ** 2 + 2.6043E-02 * float(qx) + 2.0833E+02
    return round(h2x, 5)


def qh2_w2(qx):
    h2x = -1.0596E-20 * float(qx) ** 6 + 1.5326E-16 * float(qx) ** 5 - 8.7713E-13 * float(qx) ** 4 + 2.5807E-09 * float(
        qx) ** 3 - 4.3865E-06 * float(qx) ** 2 + 6.0464E-03 * float(qx) + 1.9697E+02
    return round(h2x, 5)


def qh2_w3(qx):
    h2x = -4.0414E-14 * float(qx) ** 4 + 3.7227E-10 * float(qx) ** 3 - 1.2276E-06 * float(qx) ** 2 + 3.1791E-03 * float(
        qx) + 1.8569E+02
    return round(h2x, 5)


def qh2_w4(qx):
    h2x = -1.7856E-14 * float(qx) ** 4 + 2.1008E-10 * float(qx) ** 3 - 9.6793E-07 * float(qx) ** 2 + 3.3905E-03 * float(
        qx) + \
          1.7534E+02
    return round(h2x, 5)


def qh2_w5(qx):
    h2x = 9.8796E-18 * float(qx) ** 5 - 1.3859E-13 * float(qx) ** 4 + 7.6360E-10 * float(qx) ** 3 - 2.1600E-06 * float(
        qx) ** 2 + 4.2446E-03 * float(qx) + 1.6067E+02
    return h2x


def qh2_w6(qx):
    h2x = -4.5507E-21 * float(qx) ** 6 + 7.6477E-17 * float(qx) ** 5 - 5.1900E-13 * float(qx) ** 4 + 1.8109E-09 * float(
        qx) ** 3 - 3.4518E-06 * float(qx) ** 2 + 4.6714E-03 * float(qx) + 1.4780E+02
    return h2x


def qh2_w7(qx):
    h2x = -5.0360E-21 * float(qx) ** 6 + 8.5296E-17 * float(qx) ** 5 - 5.7766E-13 * float(qx) ** 4 + 2.0097E-09 * float(
        qx) ** 3 - 3.8819E-06 * float(qx) ** 2 + 5.2616E-03 * float(qx) + 1.3389E+02
    return h2x


def qh2_w8(qx):
    h2x = -5.1103E-21 * float(qx) ** 6 + 8.1461E-17 * float(qx) ** 5 - 5.3076E-13 * float(qx) ** 4 + 1.8394E-09 * float(
        qx) ** 3 - 3.7359E-06 * float(qx) ** 2 + 5.8813E-03 * float(qx) + 1.2257E+02
    return h2x


def qh2_w9(qx):
    h2x = -2.7608E-20 * float(qx) ** 6 + 4.0393E-16 * float(qx) ** 5 - 2.3067E-12 * float(qx) ** 4 + 6.5289E-09 * float(
        qx) ** 3 - 9.7864E-06 * float(qx) ** 2 + 9.3708E-03 * float(qx) + 1.1346E+02
    return h2x


def n_jisuan(h1x, h2x, qx):  # 计算出力结果为兆瓦
    nx = 8.5 * (float(h1x) - float(h2x)) * float(qx) / 1000
    if nx < 0:
        nx = 0
    return round(nx, 5)


# ===========================================导入函数部分结束============================================


# ================================================读取数据================================================
# 导入数据表格
wb = openpyxl.load_workbook('nsgajisuan1.xlsx')
sheet1 = wb.get_sheet_by_name('生态流量')
sheet2 = wb.get_sheet_by_name('1入库流量')
sheet3 = wb.get_sheet_by_name('区间来流丰')  # 表格名代表丰、平、枯年划分
sheet4 = wb.get_sheet_by_name('电站数据')

# 读取水库1下游生态需水量
qShengTai_w1 = np.zeros((1, 12))  # 初始化生态需水量矩阵
for i in range(2, 14):
    qShengTai_w1[[0], [i - 2]] = sheet1.cell(row=i, column=2).value  # 导入生态需水量（column=2表示最小生态流量，=3代表适宜生态流量）

# 读取水库1入库流量
qin_w1 = np.zeros((1, 12))  # 初始化入库流量矩阵
for i in range(2, 14):
    qin_w1[[0], [i - 2]] = sheet2.cell(row=i, column=2).value  # 导入水库1入库流量（column=2代表丰水年，=3代表平水年，=4代表枯水年）

# 读取区间入库流量
qQuJian_w = np.zeros((8, 12))
for j in range(3, 11):
    for i in range(2, 14):
        qQuJian_w[[j - 3], [i - 2]] = sheet3.cell(row=i, column=j).value  # 导入区间入库流量数据

z_w29 = [0 for _ in range(8)]
for i in range(8):
    z_w29[i] = sheet4.cell(row=2, column=i + 3).value  # 导入电站2-9正常蓄水位数据

n_w29 = [0 for _ in range(8)]
for i in range(8):
    n_w29[i] = sheet4.cell(row=5, column=i + 3).value  # 导入电站2-9装机容量数据

# ===============================================导入数据部分结束============================================

# 种群数量
nind = 150  # 这里设置种群数量===========================================


class MyProblem(ea.Problem):  # 继承Problem父类
    def __init__(self, M=2):
        name = 'MyProblem'  # 初始化name（函数名称，可以随意设置）
        Dim = 12  # 初始化Dim（决策变量维数）
        maxormins = [-1, 1]  # 初始化maxormins（目标最小最大化标记列表，1：最小化该目标；-1：最大化该目标）
        varTypes = [0] * Dim  # 初始化varTypes（决策变量的类型，0：实数；1：整数）
        lb = [0] * Dim  # 初始化变量下界
        for i in range(12):
            lb[i] = qin_w1[[0], [i]] - 238.8  # 决策变量下界
            if lb[i] < 0:
                lb[i] = 0

        ub = [1000] * Dim  # 初始化变量上界
        for i in range(12):
            ub[i] = np.sum(qin_w1[[0], 0:(i + 1)])  # 决策变量上界
            if ub[i] > 1000:
                ub[i] = 1000

        lbin = [1] * Dim  # 决策变量下边界（0表示不包含该变量的下边界，1表示包含）
        ubin = [1] * Dim  # 决策变量上边界（0表示不包含该变量的上边界，1表示包含）
        # 调用父类构造方法完成实例化
        ea.Problem.__init__(self, name, M, maxormins, Dim, varTypes, lb, ub, lbin, ubin)

    def aimFunc(self, pop):  # 目标函数
        Vars = pop.Phen  # 得到决策变量矩阵
        for j in range(12):
            for i in range(nind):
                if j == 0:
                    if np.sum(Vars[[i], 0:(j + 1)]) > np.sum(qin_w1[[0], 0:(j + 1)]):
                        Vars[[i], [j]] = np.sum(qin_w1[[0], 0:(j + 1)]) - 0
                    elif np.sum(Vars[[i], 0:(j + 1)]) < (np.sum(qin_w1[[0], 0:(j + 1)]) - 238.8):
                        Vars[[i], [j]] = np.sum(qin_w1[[0], 0:(j + 1)]) - 238.8

                else:
                    if np.sum(Vars[[i], 0:(j + 1)]) > np.sum(qin_w1[[0], 0:(j + 1)]):
                        Vars[[i], [j]] = np.sum(qin_w1[[0], 0:(j + 1)]) - np.sum(Vars[[i], 0:j])
                    elif np.sum(Vars[[i], 0:(j + 1)]) < (np.sum(qin_w1[[0], 0:(j + 1)]) - 238.8):
                        Vars[[i], [j]] = np.sum(qin_w1[[0], 0:(j + 1)]) - np.sum(Vars[[i], 0:j]) - 238.8

        pop.Phen = Vars

        # 修正种群符合约束
        f1 = np.zeros((nind, 1))
        f2 = np.zeros((nind, 1))
        for i in range(nind):
            v_one = np.zeros((1, 12))  # 初始化个体下泄流量过程
            v_one = Vars[[i], :]
            h1_one = np.zeros((1, 13))  # 初始化个体上游水位过程
            h1_one[[0], [0]] = 245  # 计算起始水位
            for j in range(12):
                h1_one[[0], [j + 1]] = v1h1_w1(
                    h1v1_w1(h1_one[[0], [j]]) - (v_one[[0], [j]] - qin_w1[[0], [j]]) * (730 * 3600 / 10 ** 8))
                if h1_one[[0], [j + 1]] > 275:
                    h1_one[[0], [j + 1]] = 275

            h2_one = np.zeros((1, 12))  # 初始化个体下游水位过程
            for j in range(12):
                h2_one[[0], [j]] = qh2_w1(v_one[[0], [j]])

            n_f1 = 0  # 初始化总出力记录值
            for j in range(12):
                if n_jisuan((h1_one[[0], [j]] + h1_one[[0], [j + 1]]) / 2, h2_one[[0], [j]], v_one[[0], [j]]) > 100:
                    n_f1 = n_f1 + 100
                else:
                    n_f1 = n_f1 + n_jisuan((h1_one[[0], [j]] + h1_one[[0], [j + 1]]) / 2, h2_one[[0], [j]],
                                           v_one[[0], [j]])

            n_f2 = 0  # 初始化个体电站2出力
            for j in range(8):
                qQuJian_w[[j], :] = qQuJian_w[[j], :] + v_one[[0], :]

            for k in range(8):
                for j in range(12):
                    if n_jisuan(z_w29[k], qh2_w2(qQuJian_w[[k], [j]]), qQuJian_w[[k], [j]]) < n_w29[k]:
                        n_f2 = n_f2 + n_jisuan(z_w29[k], qh2_w2(qQuJian_w[[k], [j]]), qQuJian_w[[k], [j]])
                    else:
                        n_f2 = n_f2 + n_w29[k]

            f1[[i], [0]] = (n_f1 + n_f2) * 730 * 1000  # 计算第一个目标函数 发电量kWh

            que_f2 = 0  # 初始化总缺水量记录值
            for j in range(12):
                if v_one[[0], [j]] < qShengTai_w1[[0], [j]]:
                    que_f2 = que_f2 + (qShengTai_w1[[0], [j]] - v_one[[0], [j]])

            f2[[i], [0]] = que_f2 * 730 * 3600  # 计算第二个目标函数 缺水量立方米

        # 利用可行性法则处理约束条件
        # pop.CV = np.hstack([2 - x1 - x2,
        #                     x1 + x2 - 6,
        #                     -2 - x1 + x2,
        #                     x1 - 3 * x2 - 2,
        #                     (x3 - 3) ** 2 + x4 - 4,
        #                     4 - (x5 - 3) ** 2 - x4])
        pop.ObjV = np.hstack([f1, f2])  # 把求得的目标函数值赋值给种群pop的ObjV


# -*- coding: utf-8 -*-
"""
描述:
    该案例是moea_demo1的另一个版本，展示了如何定义aimFunc()而不是evalVars()来计算目标函数和违反约束程度值。【见MyProblem.py】
    同时展示如何定义outFunc()，用于让算法在每一次进化时调用该outFunc()函数。
"""

if __name__ == '__main__':
    #  实例化问题对象
    problem = MyProblem()


    # 定义outFunc()函数
    def outFunc(alg, pop):  # alg 和 pop为outFunc的固定输入参数，分别为算法对象和每次迭代的种群对象。
        print('第 %d 代' % alg.currentGen)


    # 构建算法
    algorithm = ea.moea_NSGA2_templet(problem,
                                      ea.Population(Encoding='RI', NIND=150),
                                      MAXGEN=500,  # 最大进化代数
                                      logTras=1,  # 表示每隔多少代记录一次日志信息，0表示不记录。
                                      outFunc=outFunc)
    # 求解
    res = ea.optimize(algorithm, verbose=False, drawing=2, outputMsg=True, drawLog=True, saveFlag=False)
    print(res)

qxie_JieGuo = res.get('Vars')
f_MuBiao = res.get('ObjV')

# 将计算结果存入excel
xmq_sheet = openpyxl.Workbook()  # 创建工作表
xmq_sheet.create_sheet(index=0, title='种群决策序列')
xmq_sheet.create_sheet(index=1, title='目标函数')
sh1 = xmq_sheet.get_sheet_by_name('种群决策序列')
sh2 = xmq_sheet.get_sheet_by_name('目标函数')
for i in range(qxie_JieGuo.shape[0]):
    for j in range(qxie_JieGuo.shape[1]):
        sh1.cell(row=i+1, column=j+1).value = float(qxie_JieGuo[[i], [j]])

for i in range(f_MuBiao.shape[0]):
    for j in range(f_MuBiao.shape[1]):
        sh2.cell(row=i+1, column=j+1).value = float(f_MuBiao[[i], [j]])


print(xmq_sheet.sheetnames)
xmq_sheet.save('最少丰.xlsx')  # 保存结果